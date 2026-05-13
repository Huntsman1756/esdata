"""Load official ESMA MiFIR transaction reporting schemas.

Schema ownership is Alembic-only. This worker downloads official ESMA ZIP/XSD
files and writes deterministic schema metadata and fields.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx
import openpyxl
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).resolve().parent))

from boe import _ensure_sync_log_table, log_sync
from runtime import (
    assert_table_exists,
    configure_logging,
    ensure_database_connection,
    get_database_url,
    get_interval_seconds,
    handle_worker_failure,
    init_sentry,
)


logger = configure_logging("workers.esma_mifir_reporting")
DATABASE_URL = get_database_url()
SYNC_INTERVAL_SECONDS = get_interval_seconds("ESMA_MIFIR_REPORTING_SYNC_INTERVAL_SECONDS", 604800)

ESMA_TR_SCHEMA_PAGE = "https://www.esma.europa.eu/document/transaction-reporting-xml-schema-110"
ESMA_TR_SCHEMA_ZIP_URL = (
    "https://www.esma.europa.eu/sites/default/files/library/"
    "esma65-8-2598_annex_2_mifir_transaction_reporting_iso20022_xml_schemas.zip"
)
ESMA_TR_TECHNICAL_INSTRUCTIONS_URL = (
    "https://www.esma.europa.eu/sites/default/files/library/"
    "esma65-8-2356_mifir_transaction_reporting_technical_reporting_instructions.pdf"
)
ESMA_TR_VALIDATION_RULES_URL = (
    "https://www.esma.europa.eu/sites/default/files/library/"
    "esma65-8-2594_annex_1_mifir_transaction_reporting_validation_rules.xlsx"
)
XSD_NS = {"xs": "http://www.w3.org/2001/XMLSchema"}


@dataclass(frozen=True)
class SchemaDownload:
    source_url: str
    zip_hash: str
    files: list[tuple[str, bytes, str]]


@dataclass(frozen=True)
class ReportingDocument:
    tipo: str
    titulo: str
    referencia: str
    url_esma: str
    fecha_publicacion: str | None
    source_hash: str
    verified: bool
    completeness: str
    content: bytes


ESMA_MIFIR_DOCUMENTS: tuple[tuple[str, str, str, str, str | None, bool, str], ...] = (
    (
        "HUB",
        "ESMA MiFIR Reporting hub",
        "ESMA MiFIR Reporting",
        ESMA_TR_SCHEMA_PAGE,
        None,
        True,
        "parcial",
    ),
    (
        "SCHEMA",
        "Transaction Reporting XML Schema 1.1.0",
        "ESMA65-8-2598",
        ESMA_TR_SCHEMA_ZIP_URL,
        "2019-09-23",
        True,
        "completa",
    ),
    (
        "INSTRUCTIONS",
        "MiFIR Transaction Reporting Technical Reporting Instructions",
        "ESMA65-8-2356",
        ESMA_TR_TECHNICAL_INSTRUCTIONS_URL,
        "2018-05-29",
        True,
        "parcial",
    ),
    (
        "VALIDATION_RULES",
        "Transaction Reporting Validation Rules - 2022 update",
        "ESMA65-8-2594",
        ESMA_TR_VALIDATION_RULES_URL,
        "2022-05-31",
        True,
        "completa",
    ),
)


def fetch_schema_zip(url: str = ESMA_TR_SCHEMA_ZIP_URL) -> SchemaDownload:
    response = httpx.get(url, follow_redirects=True, timeout=60.0)
    response.raise_for_status()
    zip_hash = hashlib.md5(response.content).hexdigest()
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        files = [
            (name, archive.read(name), hashlib.md5(archive.read(name)).hexdigest())
            for name in archive.namelist()
            if name.lower().endswith(".xsd")
        ]
    if not files:
        raise RuntimeError("Official ESMA ZIP contains no XSD files")
    return SchemaDownload(source_url=str(response.url), zip_hash=zip_hash, files=files)


def fetch_reporting_documents() -> list[ReportingDocument]:
    documents: list[ReportingDocument] = []
    for tipo, titulo, referencia, url, fecha_publicacion, verified, completeness in ESMA_MIFIR_DOCUMENTS:
        response = httpx.get(url, follow_redirects=True, timeout=60.0)
        response.raise_for_status()
        documents.append(
            ReportingDocument(
                tipo=tipo,
                titulo=titulo,
                referencia=referencia,
                url_esma=str(response.url),
                fecha_publicacion=fecha_publicacion,
                source_hash=hashlib.md5(response.content).hexdigest(),
                verified=verified,
                completeness=completeness,
                content=response.content,
            )
        )
    return documents


def _local_name(value: str | None) -> str | None:
    if not value:
        return None
    return value.split(":", 1)[-1]


def _first_text(node: ET.Element | None) -> str:
    if node is None or node.text is None:
        return ""
    return " ".join(node.text.split())


def _simple_type_lengths(root: ET.Element) -> dict[str, int]:
    lengths: dict[str, int] = {}
    for simple_type in root.findall(".//xs:simpleType", XSD_NS):
        name = simple_type.get("name")
        max_length = simple_type.find(".//xs:maxLength", XSD_NS)
        if name and max_length is not None and max_length.get("value"):
            try:
                lengths[name] = int(max_length.get("value", ""))
            except ValueError:
                continue
    return lengths


def _element_length(element: ET.Element, simple_lengths: dict[str, int]) -> int | None:
    local_type = _local_name(element.get("type"))
    if local_type and local_type in simple_lengths:
        return simple_lengths[local_type]
    max_length = element.find(".//xs:maxLength", XSD_NS)
    if max_length is not None and max_length.get("value"):
        try:
            return int(max_length.get("value", ""))
        except ValueError:
            return None
    return None


def parse_xsd_fields(download: SchemaDownload) -> list[dict]:
    fields: list[dict] = []
    for filename, content, xsd_hash in download.files:
        root = ET.fromstring(content)  # noqa: S314
        simple_lengths = _simple_type_lengths(root)
        short_name = Path(filename).name
        message_name = short_name.replace(".xsd", "")
        for index, element in enumerate(root.findall(".//xs:element", XSD_NS), start=1):
            element_name = element.get("name")
            if not element_name:
                continue
            doc_node = element.find(".//xs:documentation", XSD_NS)
            min_occurs = element.get("minOccurs", "1")
            max_occurs = element.get("maxOccurs", "1")
            field_name = f"{message_name}:{index:04d}:{element_name}"
            fields.append(
                {
                    "nombre_campo": field_name[:200],
                    "tipo": (_local_name(element.get("type")) or "")[:50],
                    "longitud": _element_length(element, simple_lengths),
                    "obligatorio": min_occurs != "0",
                    "descripcion": _first_text(doc_node) or f"Official XSD element {element_name} in {short_name}",
                    "rts_referencia": None,
                    "formato": f"min={min_occurs};max={max_occurs};file={short_name}"[:100],
                    "source_url": f"{download.source_url}#{short_name}",
                    "source_hash": xsd_hash,
                    "capture_date": date.today().isoformat(),
                }
            )
    if not fields:
        raise RuntimeError("Official ESMA XSD parse returned zero fields")
    return fields


def parse_validation_rules_xlsx(document: ReportingDocument) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(document.content), data_only=True, read_only=True)
    if "TransactionDataValidations" not in workbook.sheetnames:
        raise RuntimeError("ESMA validation XLSX missing TransactionDataValidations sheet")
    worksheet = workbook["TransactionDataValidations"]
    rules: list[dict] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        codigo = str(row[0]).strip()
        field_no = "" if row[1] is None else str(row[1]).strip()
        field_name = "" if row[2] is None else str(row[2]).strip()
        validation_text = "" if row[5] is None else " ".join(str(row[5]).split())
        implementation = "" if row[6] is None else str(row[6]).strip()
        error_code = "" if row[7] is None else str(row[7]).strip()
        error_text = "" if row[8] is None else " ".join(str(row[8]).split())
        description = validation_text or error_text or f"Validation rule {codigo}"
        severity = "ERROR" if error_code or implementation in {"Application", "XML schema"} else "WARNING"
        rules.append(
            {
                "codigo": codigo[:50],
                "descripcion": description,
                "campo_afectado": f"{field_no} {field_name}".strip()[:200] or None,
                "severidad": severity,
                "rts_referencia": None,
                "source_url": document.url_esma,
                "source_hash": document.source_hash,
                "capture_date": date.today().isoformat(),
            }
        )
    if not rules:
        raise RuntimeError("ESMA validation XLSX produced zero validation rules")
    return rules


def assert_esma_schema_tables(conn) -> None:
    assert_table_exists(
        conn,
        "esma_schema",
        required_columns=("nombre", "version", "dominio", "url_esma", "source_hash", "capture_date", "verified"),
    )
    assert_table_exists(
        conn,
        "esma_schema_field",
        required_columns=("schema_id", "nombre_campo", "source_url", "source_hash", "capture_date"),
    )
    assert_table_exists(
        conn,
        "esma_reporting_document",
        required_columns=("tipo", "titulo", "referencia", "url_esma", "source_hash", "capture_date"),
    )
    assert_table_exists(
        conn,
        "esma_validation_rule",
        required_columns=("codigo", "descripcion", "source_url", "source_hash", "capture_date"),
    )


def upsert_transaction_reporting_schema(conn, download: SchemaDownload, fields: list[dict]) -> int:
    schema_id = conn.execute(
        text(
            """
            INSERT INTO esma_schema (
                nombre, version, dominio, url_esma, source_hash,
                capture_date, verified, completeness, updated_at
            )
            VALUES (
                'ESMA MiFIR Transaction Reporting ISO 20022 XML Schemas',
                '1.1.0',
                'TRANSACTION_REPORTING',
                :url_esma,
                :source_hash,
                :capture_date,
                true,
                'completa',
                now()
            )
            ON CONFLICT (nombre, version, dominio) DO UPDATE SET
                url_esma = EXCLUDED.url_esma,
                source_hash = EXCLUDED.source_hash,
                capture_date = EXCLUDED.capture_date,
                verified = EXCLUDED.verified,
                completeness = EXCLUDED.completeness,
                updated_at = now()
            RETURNING id
            """
        ),
        {
            "url_esma": download.source_url,
            "source_hash": download.zip_hash,
            "capture_date": date.today().isoformat(),
        },
    ).scalar_one()
    conn.execute(text("DELETE FROM esma_schema_field WHERE schema_id = :schema_id"), {"schema_id": schema_id})
    for field in fields:
        conn.execute(
            text(
                """
                INSERT INTO esma_schema_field (
                    schema_id, nombre_campo, tipo, longitud, obligatorio,
                    descripcion, rts_referencia, formato, source_url,
                    source_hash, capture_date, updated_at
                )
                VALUES (
                    :schema_id, :nombre_campo, :tipo, :longitud, :obligatorio,
                    :descripcion, :rts_referencia, :formato, :source_url,
                    :source_hash, :capture_date, now()
                )
                """
            ),
            {**field, "schema_id": schema_id},
        )
    return len(fields)


def upsert_reporting_documents(conn, documents: list[ReportingDocument]) -> int:
    for document in documents:
        conn.execute(text("DELETE FROM esma_reporting_document WHERE url_esma = :url_esma"), {"url_esma": document.url_esma})
    for document in documents:
        conn.execute(
            text(
                """
                INSERT INTO esma_reporting_document (
                    tipo, titulo, referencia, url_esma, fecha_publicacion,
                    source_hash, capture_date, dominio, verified,
                    completeness, updated_at
                )
                VALUES (
                    :tipo, :titulo, :referencia, :url_esma, :fecha_publicacion,
                    :source_hash, :capture_date, 'MIFIR', :verified,
                    :completeness, now()
                )
                """
            ),
            {
                "tipo": document.tipo,
                "titulo": document.titulo,
                "referencia": document.referencia,
                "url_esma": document.url_esma,
                "fecha_publicacion": date.fromisoformat(document.fecha_publicacion) if document.fecha_publicacion else None,
                "source_hash": document.source_hash,
                "capture_date": date.today().isoformat(),
                "verified": document.verified,
                "completeness": document.completeness,
            },
        )
    return len(documents)


def replace_validation_rules(conn, rules: list[dict], source_url: str) -> int:
    conn.execute(text("DELETE FROM esma_validation_rule WHERE source_url = :source_url"), {"source_url": source_url})
    for rule in rules:
        conn.execute(
            text(
                """
                INSERT INTO esma_validation_rule (
                    codigo, descripcion, campo_afectado, severidad,
                    rts_referencia, source_url, source_hash,
                    capture_date, updated_at
                )
                VALUES (
                    :codigo, :descripcion, :campo_afectado, :severidad,
                    :rts_referencia, :source_url, :source_hash,
                    :capture_date, now()
                )
                """
            ),
            rule,
        )
    return len(rules)


def run_once(worker_name: str = "worker-esma-mifir-reporting") -> dict:
    engine = create_engine(DATABASE_URL, future=True)
    ensure_database_connection(engine, logger=logger)
    sync_start = datetime.now(UTC).isoformat()
    try:
        download = fetch_schema_zip()
        fields = parse_xsd_fields(download)
        documents = fetch_reporting_documents()
        validation_document = next(document for document in documents if document.tipo == "VALIDATION_RULES")
        validation_rules = parse_validation_rules_xlsx(validation_document)
        with engine.begin() as conn:
            assert_esma_schema_tables(conn)
            field_count = upsert_transaction_reporting_schema(conn, download, fields)
            document_count = upsert_reporting_documents(conn, documents)
            rule_count = replace_validation_rules(conn, validation_rules, validation_document.url_esma)
            _ensure_sync_log_table(conn)
            log_sync(
                conn,
                worker_name,
                "ok",
                documentos_processed=document_count,
                articulos=field_count + rule_count,
                started_at=sync_start,
            )
    except Exception as exc:
        if not handle_worker_failure(engine, "esma_mifir_reporting", "TRANSACTION_REPORTING", "sync_schema", exc):
            logger.warning("ESMA MiFIR reporting sync moved to dead-letter")
        try:
            with engine.begin() as conn:
                _ensure_sync_log_table(conn)
                log_sync(conn, worker_name, "error", error_msg=str(exc)[:500], started_at=sync_start)
        except Exception as log_exc:
            logger.warning("Failed to write ESMA MiFIR reporting error log: %s", log_exc)
        raise
    return {
        "worker": worker_name,
        "files": len(download.files),
        "fields": field_count,
        "documents": document_count,
        "validation_rules": rule_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Load ESMA MiFIR transaction reporting schemas")
    parser.add_argument("--run-once", action="store_true", help="Run once and exit")
    parser.add_argument("--interval", type=int, default=None, help="Sync interval in seconds")
    args = parser.parse_args()

    init_sentry("esma_mifir_reporting")
    interval = args.interval or SYNC_INTERVAL_SECONDS
    if args.run_once:
        result = run_once()
        print(
            "[run-once] ESMA MiFIR reporting: "
            f"files={result['files']} fields={result['fields']} "
            f"documents={result['documents']} validation_rules={result['validation_rules']}"
        )
        return

    while True:
        try:
            result = run_once()
            logger.info("ESMA MiFIR reporting sync complete: %s", result)
        except Exception:
            logger.exception("Error in ESMA MiFIR reporting sync cycle")
        time.sleep(interval)


if __name__ == "__main__":
    main()
